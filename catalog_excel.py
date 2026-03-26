from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook


EXCEL_PARSER_VERSION = "catalog_excel_v1"
_ARTICLE_RE = re.compile(r"^\s*\d+\s*/\s*[A-Za-z0-9]+[A-Za-z0-9]*\s*$")
_SIZE_RE = re.compile(r"^\s*(\d{2})\s*$")


def ensure_xlsx(path: str | Path) -> Path:
    p = Path(path)
    if p.suffix.lower() != ".xls":
        return p

    out_path = p.with_suffix(".xlsx")
    xls = pd.ExcelFile(p, engine="xlrd")
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for sheet in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet, dtype=object)
            safe_name = str(sheet)[:31] if sheet else "Sheet1"
            df.to_excel(writer, sheet_name=safe_name, index=False)
    return out_path


def _norm_str(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").split()).strip()


def _as_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        try:
            return float(value)
        except Exception:
            return 0.0
    text = _norm_str(value)
    if not text:
        return 0.0
    text = text.replace(".", "").replace(",", ".") if re.search(r"\d+,\d+", text) else text
    try:
        return float(text)
    except Exception:
        return 0.0


def _find_value_after_label(row: List[Any], label: str) -> Optional[str]:
    target = label.strip().upper()
    cells = [_norm_str(cell) for cell in row]
    upper = [cell.upper() for cell in cells]
    for idx, cell in enumerate(upper):
        if cell != target:
            continue
        for jdx in range(idx + 1, min(idx + 8, len(cells))):
            value = cells[jdx]
            if value in {"", ":"}:
                continue
            return value
    return None


def _row_contains(row: List[Any], target: str) -> bool:
    needle = target.strip().upper()
    return any(_norm_str(cell).upper() == needle for cell in row)


def _update_context_from_row(row: List[Any]) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for key, label in (("fornitore", "FORNITORE"), ("reparto", "REPARTO"), ("categoria", "CATEGORIA")):
        value = _find_value_after_label(row, label)
        if value:
            out[key] = value
    if _row_contains(row, "TIPOLOGIA"):
        value = _find_value_after_label(row, "TIPOLOGIA")
        if value:
            out["tipologia"] = value
        else:
            values = [_norm_str(cell) for cell in row if _norm_str(cell)]
            if values:
                out["tipologia"] = values[-1]
    return out


def _iter_rows_from_xlsx(xlsx_path: Path, sheet: int | str = 0) -> Tuple[str, Iterable[List[Any]]]:
    workbook = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    if isinstance(sheet, int):
        idx = sheet if 0 <= sheet < len(workbook.worksheets) else 0
        worksheet = workbook.worksheets[idx]
    elif isinstance(sheet, str) and sheet.strip().isdigit():
        idx = int(sheet.strip())
        worksheet = workbook.worksheets[idx] if 0 <= idx < len(workbook.worksheets) else workbook.worksheets[0]
    else:
        worksheet = workbook[sheet] if sheet in workbook.sheetnames else workbook.worksheets[0]

    def gen():
        for row in worksheet.iter_rows(values_only=True):
            yield list(row)

    return worksheet.title, gen()


def _find_table_header(row: List[Any]) -> Optional[Dict[str, int]]:
    cells = [_norm_str(cell).upper() for cell in row]
    if not cells:
        return None

    def find_one(target: str) -> Optional[int]:
        for idx, cell in enumerate(cells):
            if cell == target:
                return idx
        return None

    neg_idx = find_one("NEG")
    giac_idx = find_one("GIAC")
    con_idx = find_one("CON")
    ven_idx = find_one("VEN")
    if neg_idx is None or giac_idx is None or con_idx is None or ven_idx is None:
        return None

    perc_idx = None
    for idx, cell in enumerate(cells):
        if cell.replace(" ", "") in {"%VEN", "PERCVEN"}:
            perc_idx = idx
            break

    size_cols: Dict[int, int] = {}
    for idx, value in enumerate(row):
        match = _SIZE_RE.match(_norm_str(value))
        if match:
            size_cols[int(match.group(1))] = idx

    return {
        "neg": neg_idx,
        "giac": giac_idx,
        "con": con_idx,
        "ven": ven_idx,
        "perc": perc_idx if perc_idx is not None else -1,
        "size_cols_json": json.dumps(size_cols, sort_keys=True),
    }


def parse_situazione_articoli_excel(
    xlsx_path: str | Path,
    sheet: int | str = 0,
) -> pd.DataFrame:
    xlsx_path = ensure_xlsx(xlsx_path)
    source_file = xlsx_path.name
    source_sheet, rows = _iter_rows_from_xlsx(xlsx_path, sheet=sheet)

    stagione_da = ""
    stagione_descr = ""
    fornitore = ""
    reparto = ""
    categoria = ""
    tipologia = ""

    table = None
    size_cols: Dict[int, int] = {}
    cur_art = ""
    cur_descr = ""
    cur_colore = ""
    out: List[Dict[str, Any]] = []

    def flush_row(neg: str, giac: Any, con: Any, ven: Any, perc: Any, row_values: List[Any]):
        if not cur_art:
            return
        neg_s = _norm_str(neg).upper()
        if neg_s in {"", "NEG", "GIAC", "CON", "VEN", "%VEN", "ARTICOLO"}:
            return
        if not re.match(r"^[A-Z0-9]{1,4}$", neg_s):
            return

        sizes: Dict[int, float] = {}
        for size, idx in size_cols.items():
            if idx >= len(row_values):
                continue
            qty = _as_float(row_values[idx])
            if qty != 0.0:
                sizes[size] = qty

        out.append(
            {
                "stagione_da": stagione_da,
                "stagione_descr": stagione_descr,
                "fornitore": fornitore,
                "reparto": reparto,
                "categoria": categoria,
                "tipologia": tipologia,
                "source_file": source_file,
                "source_sheet": source_sheet,
                "is_total": 1 if neg_s == "XX" else 0,
                "articolo": cur_art,
                "descrizione": cur_descr,
                "colore": cur_colore,
                "neg": neg_s,
                "giac": _as_float(giac),
                "con": _as_float(con),
                "ven": _as_float(ven),
                "perc_ven": _as_float(perc),
                "sizes_present": 1 if sizes else 0,
                "sizes_json": json.dumps({str(k): v for k, v in sizes.items()}, ensure_ascii=False, sort_keys=True)
                if sizes
                else "",
                "synthetic_total": "",
            }
        )

    for row in rows:
        if not row:
            continue
        if any(_norm_str(cell).upper().startswith("STAGIONE") for cell in row[:6]):
            for idx, value in enumerate(row):
                if not _norm_str(value).upper().startswith("STAGIONE"):
                    continue
                if idx + 1 < len(row):
                    stagione_da = _norm_str(row[idx + 1])
                if idx + 2 < len(row):
                    stagione_descr = _norm_str(row[idx + 2])
                break

        context = _update_context_from_row(row)
        if "fornitore" in context:
            fornitore = context["fornitore"]
        if "reparto" in context:
            reparto = context["reparto"]
        if "categoria" in context:
            categoria = context["categoria"]
        if "tipologia" in context:
            tipologia = context["tipologia"]

        header = _find_table_header(row)
        if header:
            table = header
            size_cols = json.loads(table["size_cols_json"])
            continue

        if table is None:
            continue

        art_candidate = _norm_str(row[0])
        if _ARTICLE_RE.match(art_candidate):
            cur_art = art_candidate.replace(" ", "")
            cur_descr = _norm_str(row[2]) if len(row) > 2 else ""
            cur_colore = _norm_str(row[3]) if len(row) > 3 else ""
            neg = row[table["neg"]] if table["neg"] < len(row) else ""
            giac = row[table["giac"]] if table["giac"] < len(row) else 0
            con = row[table["con"]] if table["con"] < len(row) else 0
            ven = row[table["ven"]] if table["ven"] < len(row) else 0
            perc = row[table["perc"]] if table["perc"] != -1 and table["perc"] < len(row) else 0
            flush_row(neg, giac, con, ven, perc, row)
            continue

        neg_val = row[table["neg"]] if table["neg"] < len(row) else None
        if _norm_str(neg_val) and _norm_str(neg_val).upper() != "PAGINA":
            giac = row[table["giac"]] if table["giac"] < len(row) else 0
            con = row[table["con"]] if table["con"] < len(row) else 0
            ven = row[table["ven"]] if table["ven"] < len(row) else 0
            perc = row[table["perc"]] if table["perc"] != -1 and table["perc"] < len(row) else 0
            flush_row(neg_val, giac, con, ven, perc, row)

    df = pd.DataFrame(out)
    for col in (
        "stagione_da",
        "stagione_descr",
        "fornitore",
        "reparto",
        "categoria",
        "tipologia",
        "source_file",
        "source_sheet",
        "is_total",
        "articolo",
        "descrizione",
        "colore",
        "neg",
        "giac",
        "con",
        "ven",
        "perc_ven",
        "sizes_present",
        "sizes_json",
        "synthetic_total",
    ):
        if col not in df.columns:
            df[col] = "" if col not in {"giac", "con", "ven", "perc_ven", "sizes_present", "is_total"} else 0

    for col in ("giac", "con", "ven", "perc_ven"):
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)

    key_cols = ["source_file", "source_sheet", "stagione_da", "stagione_descr", "articolo"]
    new_rows = []
    for _, group in df.groupby(key_cols, dropna=False):
        group = group.copy()
        stores = group[group["neg"] != "XX"]
        totals = group[group["neg"] == "XX"]
        sums = stores[["giac", "con", "ven"]].sum()
        con_tot = float(sums["con"])
        ven_tot = float(sums["ven"])
        perc_tot = (ven_tot / con_tot * 100.0) if con_tot else 0.0

        if not totals.empty:
            keep_idx = totals.index[0]
            df.loc[keep_idx, ["giac", "con", "ven", "perc_ven"]] = [float(sums["giac"]), con_tot, ven_tot, perc_tot]
            df.loc[keep_idx, "is_total"] = 1
            df.loc[keep_idx, "synthetic_total"] = 1
            for drop_idx in totals.index[1:].tolist():
                df = df.drop(index=drop_idx)
            continue

        base = (stores.iloc[0] if not stores.empty else group.iloc[0]).to_dict()
        base.update(
            {
                "neg": "XX",
                "is_total": 1,
                "giac": float(sums["giac"]),
                "con": con_tot,
                "ven": ven_tot,
                "perc_ven": perc_tot,
                "sizes_present": 0,
                "sizes_json": "{}",
                "synthetic_total": 1,
            }
        )
        new_rows.append(base)

    if new_rows:
        df = pd.concat([df, pd.DataFrame(new_rows)], ignore_index=True)

    return df.reset_index(drop=True)
