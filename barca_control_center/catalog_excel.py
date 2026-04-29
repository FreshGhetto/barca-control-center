from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook

from .excel_size_alignment import detect_size_data_shift


EXCEL_PARSER_VERSION = "catalog_excel_v2"
_ARTICLE_RE = re.compile(r"^\s*[A-Za-z0-9]{1,4}\s*/\s*[A-Za-z0-9]{2,}\s*$")
# Matcha: "39", "39.0", "38.5", "4.5", "42 1/2", "XXS", "XS", "S", "M", "L", "XL", "XXL", ecc.
# NOTA: la detection principale è in _parse_size_label() che è più permissiva
_SIZE_RE = re.compile(r"^\s*(\d{1,3}(?:[.,]\d)?)\s*$")
_INLINE_LABEL_CACHE: Dict[str, re.Pattern[str]] = {}


def _parse_size_label(value: Any) -> Optional[str]:
    """
    Converte un valore di cella header in una label di taglia normalizzata.

    - Interi:  39  → "39"
    - Float:   39.0 → "39"  (interi-come-float),  38.5 → "38.5"
    - Stringhe: "39", "39.0", "38,5", "38.5", "XS", "M", "42 1/2" → normalizzati
    - None / stringa vuota → None  (colonna da saltare)
    """
    if value is None:
        return None
    # Se è un numero (int/float) converti prima
    if isinstance(value, (int, float)):
        if isinstance(value, float):
            if value != value:  # NaN
                return None
            # float che è un intero (es. 39.0 → "39")
            if value == int(value):
                label = str(int(value))
            else:
                # mezze taglie (es. 38.5)
                label = str(value).replace(",", ".")
        else:
            label = str(value)
    else:
        label = " ".join(str(value).replace("\xa0", " ").split()).strip()

    if not label:
        return None

    # Normalizza separatore decimale
    label_norm = label.replace(",", ".")

    # Float-as-int in stringa: "39.0" → "39"
    try:
        f = float(label_norm)
        if f == int(f) and int(f) > 0:
            return str(int(f))
        elif f > 0:
            return label_norm  # mezza taglia: "38.5"
    except (ValueError, TypeError):
        pass

    # Taglie alfanumeriche: XS, S, M, L, XL, XXL, UNICA, ...
    upper = label.upper().strip()
    if re.match(r"^(XXS|XS|S|M|L|XL|XXL|XXXL|UNICA|TU)$", upper):
        return upper

    # Qualsiasi altro formato numerico con slash (es. "42 1/2")
    if re.match(r"^\d+\s*[/]\s*\d+$", label.strip()):
        return re.sub(r"\s+", "", label.strip())

    return None


def ensure_xlsx(path: str | Path) -> Path:
    p = Path(path)
    if p.suffix.lower() != ".xls":
        return p

    out_path = p.with_suffix(".xlsx")
    with pd.ExcelFile(p, engine="xlrd") as xls:
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            for sheet in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet, dtype=object)
                safe_name = str(sheet)[:31] if sheet else "Sheet1"
                df.to_excel(writer, sheet_name=safe_name, index=False)
    return out_path


def _norm_str(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").split()).strip()


def _as_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        try:
            return float(value)
        except Exception:
            return 0.0
    text = _norm_str(value)
    if not text:
        return 0.0
    text = text.replace(".", "").replace(",", ".") if re.search(r"\d+,\d+", text) else text
    try:
        return float(text)
    except Exception:
        return 0.0


def _find_value_after_label(row: List[Any], label: str) -> Optional[str]:
    target = label.strip().upper()
    cells = [_norm_str(cell) for cell in row]
    upper = [cell.upper() for cell in cells]
    for idx, cell in enumerate(upper):
        if cell != target:
            continue
        for jdx in range(idx + 1, min(idx + 8, len(cells))):
            value = cells[jdx]
            if value in {"", ":"}:
                continue
            return value
    pattern = _INLINE_LABEL_CACHE.get(target)
    if pattern is None:
        pattern = re.compile(rf"^\s*{re.escape(target)}\s*:?\s*(.+?)\s*$", re.IGNORECASE)
        _INLINE_LABEL_CACHE[target] = pattern
    for cell in cells:
        match = pattern.match(cell)
        if not match:
            continue
        value = _norm_str(match.group(1))
        if value and value != ":":
            return value
    return None


def _row_contains(row: List[Any], target: str) -> bool:
    needle = target.strip().upper()
    return any(_norm_str(cell).upper() == needle for cell in row)


def _update_context_from_row(row: List[Any]) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for key, label in (
        ("fornitore", "FORNITORE"),
        ("reparto", "REPARTO"),
        ("categoria", "CATEGORIA"),
        ("marchio", "MARCHIO"),
    ):
        value = _find_value_after_label(row, label)
        if value:
            out[key] = value
    if _row_contains(row, "TIPOLOGIA"):
        value = _find_value_after_label(row, "TIPOLOGIA")
        if value and value != ":":
            out["tipologia"] = value
        else:
            values = [_norm_str(cell) for cell in row if _norm_str(cell) and _norm_str(cell) != ":"]
            if values:
                out["tipologia"] = values[-1]
    return out


def _iter_rows_from_xlsx(xlsx_path: Path, sheet: int | str = 0) -> Tuple[str, Iterable[List[Any]]]:
    workbook = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    if isinstance(sheet, int):
        idx = sheet if 0 <= sheet < len(workbook.worksheets) else 0
        worksheet = workbook.worksheets[idx]
    elif isinstance(sheet, str) and sheet.strip().isdigit():
        idx = int(sheet.strip())
        worksheet = workbook.worksheets[idx] if 0 <= idx < len(workbook.worksheets) else workbook.worksheets[0]
    else:
        worksheet = workbook[sheet] if sheet in workbook.sheetnames else workbook.worksheets[0]

    def gen():
        try:
            for row in worksheet.iter_rows(values_only=True):
                yield list(row)
        finally:
            workbook.close()

    return worksheet.title, gen()


def _find_table_header(row: List[Any]) -> Optional[Dict[str, int]]:
    cells = [_norm_str(cell).upper() for cell in row]
    if not cells:
        return None

    def find_one(target: str) -> Optional[int]:
        for idx, cell in enumerate(cells):
            if cell == target:
                return idx
        return None

    neg_idx = find_one("NEG")
    giac_idx = find_one("GIAC")
    con_idx = find_one("CON")
    ven_idx = find_one("VEN")
    if neg_idx is None or giac_idx is None or con_idx is None or ven_idx is None:
        return None

    perc_idx = None
    for idx, cell in enumerate(cells):
        if cell.replace(" ", "") in {"%VEN", "PERCVEN"}:
            perc_idx = idx
            break

    # Mappa {label_taglia: col_idx} — usa _parse_size_label per gestire
    # interi, float (39.0→39, 38.5), alfanumerici (XS, M…).
    # Le colonne con header vuoto vengono saltate automaticamente (returna None).
    size_cols: Dict[str, int] = {}
    for idx, value in enumerate(row):
        label = _parse_size_label(value)
        if label is not None:
            # Se la stessa taglia appare due volte, teniamo la prima occorrenza
            if label not in size_cols:
                size_cols[label] = idx

    return {
        "neg": neg_idx,
        "giac": giac_idx,
        "con": con_idx,
        "ven": ven_idx,
        "perc": perc_idx if perc_idx is not None else -1,
        "size_cols_json": json.dumps(size_cols, sort_keys=True),
    }


def parse_situazione_articoli_excel(
    xlsx_path: str | Path,
    sheet: int | str = 0,
) -> pd.DataFrame:
    xlsx_path = ensure_xlsx(xlsx_path)
    source_file = xlsx_path.name
    source_sheet, rows = _iter_rows_from_xlsx(xlsx_path, sheet=sheet)

    stagione_da = ""
    stagione_descr = ""
    fornitore = ""
    reparto = ""
    categoria = ""
    marchio = ""
    tipologia = ""

    table = None
    size_cols: Dict[str, int] = {}  # {label_taglia: col_idx}
    cur_art = ""
    cur_descr = ""
    cur_colore = ""
    out: List[Dict[str, Any]] = []

    def flush_row(neg: str, giac: Any, con: Any, ven: Any, perc: Any, row_values: List[Any]):
        if not cur_art:
            return
        neg_s = _norm_str(neg).upper()
        if neg_s in {"", "NEG", "GIAC", "CON", "VEN", "%VEN", "ARTICOLO"}:
            return
        if not re.match(r"^[A-Z0-9]{1,4}$", neg_s):
            return

        # Legge solo le taglie presenti nell'header, saltando le colonne vuote
        sizes: Dict[str, float] = {}
        size_shift = detect_size_data_shift(row_values, size_cols, normalize=_norm_str)
        for size_label, idx in size_cols.items():
            aligned_idx = idx + size_shift
            if aligned_idx >= len(row_values):
                continue
            qty = _as_float(row_values[aligned_idx])
            sizes[size_label] = qty

        out.append(
            {
                "stagione_da": stagione_da,
                "stagione_descr": stagione_descr,
                "fornitore": fornitore,
                "reparto": reparto,
                "categoria": categoria,
                "marchio": marchio,
                "tipologia": tipologia,
                "source_file": source_file,
                "source_sheet": source_sheet,
                "is_total": 1 if neg_s == "XX" else 0,
                "articolo": cur_art,
                "descrizione": cur_descr,
                "colore": cur_colore,
                "neg": neg_s,
                "giac": _as_float(giac),
                "con": _as_float(con),
                "ven": _as_float(ven),
                "perc_ven": _as_float(perc),
                "sizes_present": 1 if size_cols else 0,
                "sizes_json": json.dumps(sizes, ensure_ascii=False, sort_keys=True)
                if size_cols
                else "",
                "synthetic_total": 0,
            }
        )

    for row in rows:
        if not row:
            continue
        if any(_norm_str(cell).upper().startswith("STAGIONE") for cell in row[:6]):
            for idx, value in enumerate(row):
                if not _norm_str(value).upper().startswith("STAGIONE"):
                    continue
                if idx + 1 < len(row):
                    stagione_da = _norm_str(row[idx + 1])
                if idx + 2 < len(row):
                    stagione_descr = _norm_str(row[idx + 2])
                break

        context = _update_context_from_row(row)
        if "fornitore" in context:
            fornitore = context["fornitore"]
        if "reparto" in context:
            reparto = context["reparto"]
        if "categoria" in context:
            categoria = context["categoria"]
        if "marchio" in context:
            marchio = context["marchio"]
        if "tipologia" in context:
            tipologia = context["tipologia"]

        header = _find_table_header(row)
        if header:
            table = header
            size_cols = json.loads(table["size_cols_json"])
            continue

        if table is None:
            continue

        art_candidate = _norm_str(row[0])
        if _ARTICLE_RE.match(art_candidate):
            cur_art = art_candidate.replace(" ", "")
            cur_descr = _norm_str(row[2]) if len(row) > 2 else ""
            cur_colore = _norm_str(row[3]) if len(row) > 3 else ""
            neg = row[table["neg"]] if table["neg"] < len(row) else ""
            giac = row[table["giac"]] if table["giac"] < len(row) else 0
            con = row[table["con"]] if table["con"] < len(row) else 0
            ven = row[table["ven"]] if table["ven"] < len(row) else 0
            perc = row[table["perc"]] if table["perc"] != -1 and table["perc"] < len(row) else 0
            flush_row(neg, giac, con, ven, perc, row)
            continue

        neg_val = row[table["neg"]] if table["neg"] < len(row) else None
        if _norm_str(neg_val) and _norm_str(neg_val).upper() != "PAGINA":
            giac = row[table["giac"]] if table["giac"] < len(row) else 0
            con = row[table["con"]] if table["con"] < len(row) else 0
            ven = row[table["ven"]] if table["ven"] < len(row) else 0
            perc = row[table["perc"]] if table["perc"] != -1 and table["perc"] < len(row) else 0
            flush_row(neg_val, giac, con, ven, perc, row)

    df = pd.DataFrame(out)
    for col in (
        "stagione_da",
        "stagione_descr",
        "fornitore",
        "reparto",
        "categoria",
        "marchio",
        "tipologia",
        "source_file",
        "source_sheet",
        "is_total",
        "articolo",
        "descrizione",
        "colore",
        "neg",
        "giac",
        "con",
        "ven",
        "perc_ven",
        "sizes_present",
        "sizes_json",
        "synthetic_total",
    ):
        if col not in df.columns:
            df[col] = "" if col not in {"giac", "con", "ven", "perc_ven", "sizes_present", "is_total"} else 0

    for col in ("giac", "con", "ven", "perc_ven"):
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)

    key_cols = ["source_file", "source_sheet", "stagione_da", "stagione_descr", "articolo"]
    new_rows = []
    for _, group in df.groupby(key_cols, dropna=False):
        group = group.copy()
        stores = group[group["neg"] != "XX"]
        totals = group[group["neg"] == "XX"]
        sums = stores[["giac", "con", "ven"]].sum()
        con_tot = float(sums["con"])
        ven_tot = float(sums["ven"])
        perc_tot = (ven_tot / con_tot * 100.0) if con_tot else 0.0

        if not totals.empty:
            keep_idx = totals.index[0]
            df.loc[keep_idx, ["giac", "con", "ven", "perc_ven"]] = [float(sums["giac"]), con_tot, ven_tot, perc_tot]
            df.loc[keep_idx, "is_total"] = 1
            df.loc[keep_idx, "synthetic_total"] = 1
            for drop_idx in totals.index[1:].tolist():
                df = df.drop(index=drop_idx)
            continue

        base = (stores.iloc[0] if not stores.empty else group.iloc[0]).to_dict()
        base.update(
            {
                "neg": "XX",
                "is_total": 1,
                "giac": float(sums["giac"]),
                "con": con_tot,
                "ven": ven_tot,
                "perc_ven": perc_tot,
                "sizes_present": 0,
                "sizes_json": "{}",
                "synthetic_total": 1,
            }
        )
        new_rows.append(base)

    if new_rows:
        df = pd.concat([df, pd.DataFrame(new_rows)], ignore_index=True)

    return df.reset_index(drop=True)
