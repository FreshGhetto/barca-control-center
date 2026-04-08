from __future__ import annotations

import argparse
import csv
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional

import pandas as pd
import xlrd

try:
    import openpyxl
except Exception:  # pragma: no cover
    openpyxl = None

from .db_sync import run_db_sync
from .pipeline_common import harmonize_clean_frames, load_valid_shop_codes
from .reparto_sizes import SUPPORTED_SIZES, infer_reparto_from_path


ARTICLE_CODE_RE = re.compile(r"^\d{1,3}/\S+")
SEASON_RE = re.compile(r"(?P<season>\d{2}[a-z])", re.IGNORECASE)
SHOP_ALIASES = {
    "W": "WEB",
    "NU": "NV",
    "M2": "ME2",
}
SIZE_LABEL_RE = re.compile(r"^\d{2}$")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Import storico BARCA da report negozio .xls/.xlsx/.csv con righe articolo-negozio."
    )
    parser.add_argument(
        "--stock-dir",
        type=Path,
        required=True,
        help="Cartella con i report shop-level (es. ventite_per_negozio_uomo).",
    )
    parser.add_argument(
        "--db-create-schema",
        action="store_true",
        help="Applica db/schema.sql prima del primo import.",
    )
    parser.add_argument(
        "--season-filter",
        nargs="*",
        default=None,
        help="Lista stagioni opzionale da importare (es. 23i 24y 25i).",
    )
    return parser.parse_args()


def _season_sort_key(code: str):
    code = str(code or "").strip().lower()
    year = int(code[:2]) + 2000 if len(code) >= 2 and code[:2].isdigit() else -1
    season_char = code[-1] if code else ""
    season_rank = {"y": 0, "g": 0, "i": 1, "e": 1}.get(season_char, 9)
    return year, season_rank, code


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip()
    return text


def _clean_number(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        try:
            return float(value)
        except Exception:
            return 0.0
    text = _clean_text(value)
    if not text or text == "-":
        return 0.0
    text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except Exception:
        return 0.0


def _clean_non_negative(value: object) -> float:
    return max(0.0, _clean_number(value))


def _normalize_shop_code(value: object) -> str:
    code = _clean_text(value).split(" ")[0].upper()
    return SHOP_ALIASES.get(code, code)


def _is_article_code(value: object) -> bool:
    text = _clean_text(value)
    if not text:
        return False
    return bool(ARTICLE_CODE_RE.match(text.split(" ")[0]))


def _iter_rows(path: Path) -> Iterable[List[str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with open(path, "r", encoding="latin1", errors="ignore") as handle:
            reader = csv.reader(handle)
            for row in reader:
                yield [_clean_text(cell) for cell in row]
        return

    if suffix == ".xls":
        book = xlrd.open_workbook(path)
        try:
            sheet = book.sheet_by_index(0)
            for ridx in range(sheet.nrows):
                yield [_clean_text(sheet.cell_value(ridx, cidx)) for cidx in range(sheet.ncols)]
        finally:
            book.release_resources()
        return

    if suffix in {".xlsx", ".xlsm"}:
        if openpyxl is None:
            raise RuntimeError("openpyxl non disponibile per leggere file .xlsx/.xlsm")
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.worksheets[0]
            for row in sheet.iter_rows(values_only=True):
                yield [_clean_text(cell) for cell in row]
        finally:
            workbook.close()
        return

    raise ValueError(f"Formato non supportato: {path}")


def _infer_season_code(path: Path) -> str:
    match = SEASON_RE.search(path.stem)
    if not match:
        raise ValueError(f"Stagione non rilevata dal nome file: {path.name}")
    return match.group("season").lower()


def _snapshot_at_for_file(path: Path) -> str:
    return datetime.fromtimestamp(path.stat().st_mtime).isoformat()


def parse_shop_stock_report(path: Path, valid_codes: Optional[List[str]] = None) -> pd.DataFrame:
    valid_set = {str(code).strip().upper() for code in (valid_codes or []) if str(code).strip()}
    size_map: Dict[int, int] = {}
    current_article = ""
    current_desc = ""
    reparto = infer_reparto_from_path(path) or ""
    rows: List[Dict[str, object]] = []

    for row in _iter_rows(path):
        if not any(row):
            continue

        upper_row = [_clean_text(cell).upper() for cell in row]
        if "NEG" in upper_row and "GIAC" in upper_row and "VEN" in upper_row:
            size_map = {}
            for idx, cell in enumerate(upper_row):
                label = cell.strip()
                if SIZE_LABEL_RE.match(label):
                    size_map[idx] = int(label)
            continue

        if _is_article_code(row[0] if row else ""):
            current_article = _clean_text(row[0]).split(" ")[0].upper()
            desc_parts = []
            if len(row) > 2 and _clean_text(row[2]):
                desc_parts.append(_clean_text(row[2]))
            if len(row) > 3 and _clean_text(row[3]):
                desc_parts.append(_clean_text(row[3]))
            current_desc = " ".join(desc_parts).strip()

        if not current_article:
            continue

        shop = _normalize_shop_code(row[4] if len(row) > 4 else "")
        if not shop:
            continue
        if valid_set and shop not in valid_set:
            continue

        size_values = {size: 0.0 for size in SUPPORTED_SIZES}
        for idx, size in size_map.items():
            if size in size_values and idx < len(row):
                size_values[size] = _clean_non_negative(row[idx])

        record = {
            "Article": current_article,
            "Description": current_desc,
            "Reparto": reparto,
            "Shop": shop,
            "Ricevuto": 0.0,
            "Giacenza": _clean_non_negative(row[5] if len(row) > 5 else 0.0),
            "Consegnato": _clean_non_negative(row[6] if len(row) > 6 else 0.0),
            "Venduto": _clean_non_negative(row[7] if len(row) > 7 else 0.0),
            "Sellout_Percent": _clean_non_negative(row[8] if len(row) > 8 else 0.0),
            "Valore_Giac": 0.0,
        }
        for size in SUPPORTED_SIZES:
            record[f"Size_{size}"] = size_values[size]
        rows.append(record)

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    return df.drop_duplicates(subset=["Article", "Shop"], keep="last").reset_index(drop=True)


def build_sales_from_stock(stock_df: pd.DataFrame) -> pd.DataFrame:
    if stock_df.empty:
        return pd.DataFrame(
            columns=[
                "snapshot_at",
                "Article",
                "Shop",
                "Consegnato_Qty",
                "Venduto_Qty",
                "Periodo_Qty",
                "Altro_Venduto_Qty",
                "Sellout_Percent",
                "Sellout_Clamped",
                "Valore_1",
                "Valore_2",
                "Valore_3",
                "Valore_4",
            ]
        )

    sales = pd.DataFrame(
        {
            "snapshot_at": stock_df["snapshot_at"],
            "Article": stock_df["Article"],
            "Shop": stock_df["Shop"],
            "Consegnato_Qty": pd.to_numeric(stock_df["Consegnato"], errors="coerce").fillna(0.0).clip(lower=0.0),
            "Venduto_Qty": pd.to_numeric(stock_df["Venduto"], errors="coerce").fillna(0.0).clip(lower=0.0),
            "Periodo_Qty": pd.to_numeric(stock_df["Venduto"], errors="coerce").fillna(0.0).clip(lower=0.0),
            "Altro_Venduto_Qty": 0.0,
            "Sellout_Percent": pd.to_numeric(stock_df["Sellout_Percent"], errors="coerce").fillna(0.0).clip(lower=0.0),
            "Sellout_Clamped": pd.to_numeric(stock_df["Sellout_Percent"], errors="coerce").fillna(0.0).clip(lower=0.0, upper=100.0),
            "Valore_1": 0.0,
            "Valore_2": 0.0,
            "Valore_3": 0.0,
            "Valore_4": 0.0,
        }
    )
    return sales.drop_duplicates(subset=["Article", "Shop"], keep="last").reset_index(drop=True)


def main():
    args = parse_args()
    root = Path(__file__).resolve().parent.parent
    stock_dir = args.stock_dir.resolve()
    if not stock_dir.exists():
        raise SystemExit(f"Cartella non trovata: {stock_dir}")

    shops_cfg = root / "config" / "lista-negozi_integrato.xlsx"
    if not shops_cfg.exists():
        shops_cfg = root / "config" / "lista-negozi.xlsx"
    valid_shop_codes = load_valid_shop_codes(shops_cfg)

    files = [p for p in stock_dir.iterdir() if p.is_file() and p.suffix.lower() in {".xls", ".xlsx", ".xlsm", ".csv"}]
    if args.season_filter:
        wanted = {str(item).strip().lower() for item in args.season_filter if str(item).strip()}
        files = [p for p in files if _infer_season_code(p) in wanted]

    season_to_file = {}
    for path in files:
        try:
            season_to_file[_infer_season_code(path)] = path
        except Exception:
            continue

    ordered = sorted(season_to_file.items(), key=lambda item: _season_sort_key(item[0]))
    if not ordered:
        raise SystemExit("Nessun report stagionale importabile trovato.")

    history_dir = root / "output" / "history_import"
    history_dir.mkdir(parents=True, exist_ok=True)
    manifest_rows: List[Dict[str, object]] = []
    latest_raw_run_id = None

    for idx, (season_code, path) in enumerate(ordered, start=1):
        print(f"[{idx}/{len(ordered)}] Import {season_code} da {path.name}")
        snapshot_at = _snapshot_at_for_file(path)
        stock_df = parse_shop_stock_report(path, valid_codes=valid_shop_codes)
        if stock_df.empty:
            print(f"  - saltato: nessuna riga articolo/negozio estratta da {path.name}")
            manifest_rows.append(
                {
                    "season_code": season_code,
                    "source_file": str(path),
                    "status": "skipped",
                    "reason": "no_rows",
                }
            )
            continue

        stock_df = stock_df.copy()
        stock_df.insert(0, "snapshot_at", snapshot_at)
        sales_df = build_sales_from_stock(stock_df)
        sales_df, stock_df, align_report = harmonize_clean_frames(sales_df, stock_df)

        season_dir = history_dir / season_code
        season_dir.mkdir(parents=True, exist_ok=True)
        sales_df.to_csv(season_dir / f"clean_sales_{season_code}.csv", index=False)
        stock_df.to_csv(season_dir / f"clean_stock_{season_code}.csv", index=False)
        align_report.to_csv(season_dir / f"alignment_{season_code}.csv", index=False)

        summary = run_db_sync(
            root=root,
            create_schema=bool(args.db_create_schema and latest_raw_run_id is None),
            run_type="raw_input_sync",
            verbose=True,
            clean_sales_df=sales_df,
            clean_stock_df=stock_df,
            transfers_df=pd.DataFrame(),
            features_df=pd.DataFrame(),
            ingest_report={
                "source_mode": "shop_report_history",
                "source_file": str(path),
                "season_code": season_code,
            },
            include_orders=False,
            metadata_extra={
                "operating_mode": "raw_to_db",
                "source_path_mode": "shop_report_history",
                "season_code": season_code,
                "snapshot_at": snapshot_at,
                "source_file": str(path),
                "source_kind": "shop_stock_report",
                "sales_basis": "derived_from_shop_stock_report",
            },
        )
        latest_raw_run_id = summary.get("run_id")
        manifest_rows.append(
            {
                "season_code": season_code,
                "source_file": str(path),
                "status": "imported",
                "snapshot_at": snapshot_at,
                "sales_rows": int(len(sales_df)),
                "stock_rows": int(len(stock_df)),
                "alignment_rows": int(len(align_report)),
                "run_id": latest_raw_run_id,
            }
        )

    manifest = pd.DataFrame(manifest_rows)
    manifest_path = history_dir / "shop_report_import_manifest.csv"
    manifest.to_csv(manifest_path, index=False)
    print(f"\nManifest scritto in {manifest_path}")
    if latest_raw_run_id:
        print(f"Ultimo raw_input_sync importato: {latest_raw_run_id}")
    else:
        print("Nessun run raw_input_sync creato.")


if __name__ == "__main__":
    main()
