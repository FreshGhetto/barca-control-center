from __future__ import annotations

import csv
import datetime as dt
import json
import re
import shutil
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl


SUPPORTED_EXT = {".csv", ".xlsx", ".xlsm", ".xls"}


def _now_tag() -> str:
    return dt.datetime.now().strftime("%Y%m%d_%H%M%S")


def _ensure_dir(path: Path):
    path.mkdir(parents=True, exist_ok=True)


def _safe_name(name: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", name).strip("_")


def _extract_report_date(text: str) -> Optional[dt.date]:
    m = re.search(r"(\d{2})[./-](\d{2})[./-](\d{4})", text)
    if not m:
        return None
    day, month, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
    try:
        return dt.date(year, month, day)
    except Exception:
        return None


def _extract_season_code(text: str, filename: str) -> Optional[str]:
    t = text.upper()
    m = re.search(r"STAGIONE[:\s]*([0-9]{2}[A-Z])", t)
    if m:
        return m.group(1).lower()
    m = re.search(r"([0-9]{2}[a-zA-Z])_sd_[1-4]", filename, flags=re.IGNORECASE)
    if m:
        return m.group(1).lower()
    return None


def _read_csv_preview(path: Path, max_lines: int = 120) -> str:
    encodings = ("utf-8", "utf-8-sig", "cp1252", "latin-1")
    for enc in encodings:
        try:
            with open(path, "r", encoding=enc, errors="replace") as f:
                lines = []
                for idx, line in enumerate(f):
                    if idx >= max_lines:
                        break
                    lines.append(line.rstrip("\n"))
                return "\n".join(lines)
        except Exception:
            continue
    return ""


def _read_xlsx_preview(path: Path, max_rows: int = 120, max_cols: int = 120) -> str:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        rows = []
        for ridx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if ridx > max_rows:
                break
            vals = []
            for cidx, val in enumerate(row, start=1):
                if cidx > max_cols:
                    break
                vals.append("" if val is None else str(val))
            rows.append(",".join(vals))
        return "\n".join(rows)
    finally:
        wb.close()


def _read_preview(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv_preview(path)
    if suffix in {".xlsx", ".xlsm", ".xls"}:
        return _read_xlsx_preview(path)
    return ""


def _classify_file(path: Path, preview: str) -> Dict[str, Any]:
    text = preview.upper()
    fname = path.name.upper()

    if "SITUAZIONE ARTICOLI PER NEGOZIO" in text:
        return {"kind": "stock_report", "target_group": "distribution"}

    if "ANALISI LISTINI E RICARICHI" in text:
        return {"kind": "orders_prices", "target_group": "orders"}

    if "_SD_3" in fname or "ANALISI PER SINGOLA TAGLIA" in text:
        return {"kind": "orders_sd_3", "target_group": "orders"}

    if "_SD_4" in fname or ("FASCE PRZ." in text and "LISTINO" in text and "ANALISI ARTICOLI" in text):
        return {"kind": "orders_sd_4", "target_group": "orders"}

    if "_SD_1" in fname:
        return {"kind": "orders_sd_1", "target_group": "orders"}

    if "_SD_2" in fname:
        return {"kind": "orders_sd_2", "target_group": "orders"}

    if "ANALISI ARTICOLI" in text:
        if "NEGOZIO" in text and "FORNITORE" in text:
            return {"kind": "sales_report", "target_group": "distribution"}
        if "TIPOLOGIA" in text and "MARCHIO" in text:
            return {"kind": "orders_sd_1", "target_group": "orders"}
        if "COLORE" in text and "MATERIALE" in text:
            return {"kind": "orders_sd_2", "target_group": "orders"}
        if "TAG" in text and "TOT" in text:
            return {"kind": "orders_sd_3", "target_group": "orders"}

    return {"kind": "unknown", "target_group": "quarantine"}


def _convert_to_csv(src: Path, dst: Path):
    _ensure_dir(dst.parent)
    suffix = src.suffix.lower()
    if suffix == ".csv":
        shutil.copy2(src, dst)
        return

    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        with open(dst, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(["" if v is None else str(v) for v in row])
    finally:
        wb.close()


def _target_path(root: Path, kind: str, report_date: Optional[dt.date], season_code: Optional[str]) -> Optional[Path]:
    month_tag = (report_date or dt.date.today()).strftime("%Y-%m")
    if kind == "sales_report":
        return root / "input" / f"sales_{month_tag}.csv"
    if kind == "stock_report":
        return root / "input" / f"stock_{month_tag}.csv"

    if kind.startswith("orders_"):
        orders_dir = root / "input" / "orders"
        if kind == "orders_prices":
            if season_code:
                return orders_dir / f"{season_code}_prezzo_acq-ven.csv"
            return orders_dir / "prezzo_acq-ven.csv"
        if kind in {"orders_sd_1", "orders_sd_2", "orders_sd_3", "orders_sd_4"}:
            sd_num = kind.rsplit("_", 1)[-1]
            if not season_code:
                return orders_dir / f"unknown_sd_{sd_num}.csv"
            return orders_dir / f"{season_code}_sd_{sd_num}.csv"
    return None


def ingest_incoming(
    root: Path,
    incoming_dir: Optional[Path] = None,
    *,
    move_processed: bool = True,
    verbose: bool = True,
) -> Dict[str, Any]:
    incoming = incoming_dir or (root / "incoming")
    quarantine_dir = incoming / "_quarantine"
    archive_dir = incoming / "_processed"
    reports_dir = root / "output" / "ingest"
    _ensure_dir(incoming)
    _ensure_dir(quarantine_dir)
    _ensure_dir(archive_dir)
    _ensure_dir(reports_dir)

    files = [
        p
        for p in incoming.rglob("*")
        if p.is_file()
        and p.suffix.lower() in SUPPORTED_EXT
        and "_quarantine" not in p.parts
        and "_processed" not in p.parts
    ]
    files.sort(key=lambda p: p.stat().st_mtime)

    rows: List[Dict[str, Any]] = []
    for src in files:
        rec: Dict[str, Any] = {
            "source": str(src),
            "filename": src.name,
            "status": "unknown",
            "kind": "",
            "target": "",
            "note": "",
        }
        try:
            preview = _read_preview(src)
            cls = _classify_file(src, preview)
            kind = cls["kind"]
            rec["kind"] = kind

            report_date = _extract_report_date(preview)
            season_code = _extract_season_code(preview, src.name)
            target = _target_path(root, kind, report_date, season_code)

            if kind == "unknown" or target is None:
                q_name = f"{_now_tag()}__{_safe_name(src.name)}"
                q_path = quarantine_dir / q_name
                shutil.copy2(src, q_path)
                rec["status"] = "quarantine"
                rec["target"] = str(q_path)
                rec["note"] = "Formato non riconosciuto"
                if move_processed:
                    src.unlink(missing_ok=True)
                rows.append(rec)
                if verbose:
                    print(f"[INGEST] QUARANTINE {src.name}")
                continue

            _convert_to_csv(src, target)
            rec["status"] = "ingested"
            rec["target"] = str(target)
            rec["note"] = f"season={season_code or ''} date={(report_date.isoformat() if report_date else '')}"

            if move_processed:
                arc_name = f"{_now_tag()}__{_safe_name(src.name)}"
                arc_path = archive_dir / arc_name
                shutil.move(str(src), str(arc_path))

            if verbose:
                print(f"[INGEST] OK {src.name} -> {target.relative_to(root)} ({kind})")
        except Exception as exc:
            rec["status"] = "error"
            rec["note"] = str(exc)
            if verbose:
                print(f"[INGEST] ERROR {src.name}: {exc}")
        rows.append(rec)

    summary = {
        "timestamp": dt.datetime.now().isoformat(timespec="seconds"),
        "incoming_dir": str(incoming),
        "processed_total": len(rows),
        "ingested": int(sum(1 for r in rows if r["status"] == "ingested")),
        "quarantine": int(sum(1 for r in rows if r["status"] == "quarantine")),
        "errors": int(sum(1 for r in rows if r["status"] == "error")),
        "rows": rows,
    }

    tag = _now_tag()
    (reports_dir / "ingest_report_latest.json").write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
    (reports_dir / f"ingest_report_{tag}.json").write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
    with open(reports_dir / "ingest_report_latest.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["source", "filename", "status", "kind", "target", "note"])
        writer.writeheader()
        writer.writerows(rows)

    return summary

